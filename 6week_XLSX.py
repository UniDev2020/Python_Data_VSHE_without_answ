#XLSX
from openpyxl import load_workbook

wb = load_workbook('test.xlsx')                              # считать файл xlsx
print(wb.sheetnames)                                         # показать название листов

sheet = wb['Лист1']                                          # работа с листом
print(sheet.title)                                           # напечатать название листа
print(sheet['A1'].value)                                     # вывод значения ячейки A1

c = sheet['B2']
print(c.column)                                              # вывести номер столбца
print(c.row)                                                 # вывести номер строки
print(c.coordinate)                                          # вывести координату

d = sheet.cell(row=1, column=2)                              # определяем значение по координатам
print(d.value)
print("#---")
print(sheet.max_row)
print(sheet.max_column)
print("-#1-")
for i in range(1, sheet.max_row + 1):                        # проходим по всем строкам
    for j in range(1, sheet.max_column + 1):                 # проходим по всем колонкам
        d = sheet.cell(row=i, column=j)                      # определяем значение по координатам во всей таблице
        print(d.value)
print("-#2-")
for cellObj in sheet['A1': 'C3']:
    for cell in cellObj:
        print(cell.value)
print('#---')
