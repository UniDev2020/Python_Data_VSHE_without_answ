'''
from openpyxl import load_workbook

salary = {}
regions = {}

wb = load_workbook('test2.xlsx')
sheet = wb['Лист1']
sheet2 = wb['Лист2']

for i in range(1, sheet.max_row + 1):                          # Работа с первым листом книги
    name = sheet.cell(row=i, column=1).value
    sal = sheet.cell(row=i, column=2).value
    salary[name] = sal
print(salary)

for i in range(1, sheet2.max_row + 1):                         # Работа с вторым листом книги
    name = sheet2.cell(row=i, column=1).value
    region = sheet2.cell(row=i, column=2).value
    if region not in regions:
        regions[region] = [0, 0]
    regions[region][0] += 1
    regions[region][1] += salary[name]
for region in regions:
    print(region, regions[region][1]/regions[region][0])
print(#---)
'''
from openpyxl import load_workbook
salary = {}
wb = load_workbook('test4.xlsx')
sheet = wb['Лист1']

for i in range(1, sheet.max_row + 1):                          # Работа с первым листом книги
    name = sheet.cell(row=i, column=1).value
    sal = sheet.cell(row=i, column=2).value
    salary[name] = sal
print(salary)
sheet.append(['Sasha', 4])
wb.save('test4.xlsx')
